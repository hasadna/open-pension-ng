#from django.db import models
#from django.core.management import settings
#from django.core.management import execute_from_command_line
#from pathlib import Path
#BASE_DIR = Path(__file__).resolve().parent.parent
#settings.configure(DEBUG=False,
#DATABASES={'default':{
#	'ENGINE':'django.db.backends.sqlite3',
#	'NAME':BASE_DIR / "db.sqlite3",
#}
#},
#INSTALLED_APPS=('importer',)
#)


#from importer.models import *
import openpyxl
from openpyxl import load_workbook
import os
import datetime
import traceback
import argparse
import json

MAPPING_FILE = "/home/guyga/hasadna/penssion/open-pension-next-generation/open_pension/field_mapping.json"

class xls_ingester(object):
    file = None
    wb = None
    with open(MAPPING_FILE) as json_data:
            mapping = json.load(json_data)

    def __init__(self):
        super().__init__()
        
	
    def find_sn(self, wb):
        sheet_name_array = wb.sheetnames
        for sn in sheet_name_array:
            # ignore dynamic pick lists that exists in some of the reports
            if sn.startswith("{PL}PickLst"):
                sheet_name_array.remove(sn)           
        for tab in self.mapping:
            sn = tab["tab_name"]
            if sn != "any":
                self.parse_first_tab(wb, sn, tab)
                sheet_name_array.remove(sn)
                ws = wb[sn]
                fields = tab["fields"]
                for field in fields:
               	    if field["type"] == "generated":
                        continue
                        #print(field["column_title"])   
            #print(tab)
        return sheet_name_array, tab

    def find_title_row(self, ws, field_name):
     	#print(field_name)
        for row in ws.iter_rows():
            for cell1 in row:
                if cell1.value is not None:
                    if cell1.value in field_name:
                   		#print("cell="+cell1.value)
                        return cell1.row
                
    def get_value_for(self, ws, field_name):
        for row in ws.iter_rows():
            for cell in row:
                if cell1.value is not None:
               	    if cell.value == field_name:
                        if ws.cell(row=cell1.row+1,column=cell1.column).value is not None:
                        	#print(ws.cell(row=cell1.row+1,column=cell1.column).value)
                            return ws.cell(row=cell1.row+1,column=cell1.column).value

    def parse_first_tab(self, wb, sn, tab):
		# from first tab get report date, company, track name and trach code
		# optinally get report summary as well
        columns = values = ""
        for field in tab["fields"]:
            #print(field)
            if field["type"] ==  'generated':
                continue
            if field["type"] ==  'extracted':
                for row in wb[sn].iter_rows(min_row=1, max_row=4, max_col=4,values_only=False):	
                   for cell in row:
                    found = False
                    if cell.value is not None:
                    	#print("========"+str(cell.value))
                   		#print(titles)
                        if cell.value.startswith("{PL}PickLst"):
                            break
                        stripped = cell.value.replace('*','').replace(":","").strip()
                    	#if stripped != cell.value :
                        #print("stripped="+stripped+"cell_value=~"+cell.value+"~")    
                        for field in tab["fields"]:
                       		#in some of the reports there are multiple * characters of column titles as pointers to comments
                            if stripped in field["column_title"]:
                                found = True
                                i = 1
                                for i in range(1,4):
                                    val = wb[sn].cell(row=cell.row,column=cell.column+i).value
                                    if val is not None:
                                        columns = columns+","+field["column_name"]
                                        values = values+",\""+str(val)+"\""
                                        break
                                break
                        break
                return
#here need to iterate cells of row to get the value



    def parse_spreadsheet(self, wb):
        sheet_name_array , tab = self.find_sn(wb)
        for sh in sheet_name_array:
            titles = list()
            column_idxs = list()
       	    column_list = list()
            columns = values = ""
            title_row = 0
        	#print(sh)
            for field in tab["fields"]:
            	#print(field)
                if field["type"] ==  'generated':
                    continue
                if field["type"] ==  'extracted':
                    title_row = self.find_title_row(wb[sh], field["column_title"])
                    if title_row is not None:
						#print("title_row="+str(title_row)+","+"column_title="+str(field["column_title"]))
                        break
                	    #columns = columns + field["column_name"]+","
                	    #values = values + str(get_value_for(wb[sh], field["column_title"]))+","
        		#print(columns+',\n\r'+values)
            for row in wb[sh].iter_rows(min_row=title_row, max_row=title_row,values_only=False):
                for cell in row:
                    found = False
                    if cell.value is not None:
                    	#print("========"+str(cell.value))
                   		#print(titles)
                        if cell.value.startswith("{PL}PickLst"):
                            break
                        stripped = cell.value.replace('*','').strip()
                    	#if stripped != cell.value :
                        #print("stripped="+stripped+"cell_value=~"+cell.value+"~")    
                        for field in tab["fields"]:
                       		#in some of the reports there are multiple * characters of column titles as pointers to comments
                            if stripped in field["column_title"]:
                                found = True
                                titles.append(cell.value)
                                column_idxs.append(cell.column)
                                column_list.append(field["column_name"])
                                break
                       		#else:
                            	#print("stripped="+stripped)
                        if not found:    
                      		# print("stripped="+stripped+", fileds="+str(tab["fields"]))
                            with open("notfound_list.txt", 'a') as fileOUT:
                                fileOUT.write(sh +"-"+str(cell.value)+"- not found in metadata"+"\n\r")
                                fileOUT.close()
                        		#print(str(cell.value)+"- not found in metadata")
                    	#print(column_idxs) 
            for c in column_list:
                columns = columns + ","+c
        		#columns = str(column_list)
            done = False            
            for row in wb[sh].iter_rows(min_row=title_row+1):
                if row[column_idxs[0]-1].value is None:
                	#print(str(row))
                    continue
                for i in range(len(column_list)):
                    value = str(row[column_idxs[i]-1].value)
                    if '*' in value:
                        done = True    
                    else:   
                        if value == "None":
                            value = "Null"
                        else:
                            value = "\""+value+"\""
                        values = values +","+value
                    if done:
                        break
                with open("inserts.sql", 'a') as fileOUT:
                    fileOUT.write(str(len(column_list))+"\n\r")
                    fileOUT.write("insert into ASSET_DETAILS ("+columns[1:]+")"+"\n\r values ("+values[1:]+");\n\r")
                    fileOUT.close()    
                values = ""    
                if done: 
           	        break
           		# print("insert into ASSET_DETAILS ("+columns[1:len(columns)-1]+")")
           		# print("values ("+values[1:]+")")
                values = ""
    
    def ingest(self, file):
        try:
            wb = load_workbook(file)
            self.parse_spreadsheet(wb)
        except Exception as e:
            traceback.print_exc()
            with open("failed_list.txt", 'a') as fileOUT:
                fileOUT.write(file+"\n\r")
                fileOUT.close()			

def xls_import(path, file):
    if path is None:
        ingester = xls_ingester()
        ingester.ingest(file)
    else:
        files = os.listdir(path)
        files = [os.path.join(path, f) for f in files if not f.startswith(".")]
        for f in files:
            ingester = xls_ingester()
            ingester.ingest(f)


def main():  # Main
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--full_file_name', type=str)
    parser.add_argument('-d', '--path', type=str)
    args = parser.parse_args()
    if args.path is None and args.full_file_name is None:
        print("specify either single file or directory")
        return
    if args.path is not None and args.full_file_name is not None:
        print("specify either single file or directory")
        return
    xls_import(args.path, args.full_file_name)


if __name__ == '__main__':
    main()           
