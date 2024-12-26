import importer
from importer import models
import openpyxl
from openpyxl import load_workbook
import os
import datetime
import traceback
import argparse
import json
import datetime
from dateutil import parser
from efc.interfaces.iopenpyxl import OpenpyxlInterface
from pycel import ExcelCompiler

MAPPING_FILE = "./importer/field_mapping.json"


class xls_ingester(object):
    force = True
    file = None
    wb = None
    with open(MAPPING_FILE) as json_data:
        mapping = json.load(json_data)
    reference_objects = dict()

    def __init__(self):
        super().__init__()

    def get_sheetnames(self, wb):
        return wb.sheetnames

    def find_sn(self, wb):
        sheet_name_array = self.get_sheetnames(wb)
        for sn in sheet_name_array:
            # ignore dynamic pick lists that exists in some of the reports
            if sn.startswith("{PL}PickLst"):
                sheet_name_array.remove(sn)
        for tab in self.mapping:
            sn = tab["tab_name"]
            if sn != "any":
                sn2 = self.parse_first_tab(wb, sn, tab)
                sheet_name_array.remove(sn2)
                #ws = wb[sn2]
                #fields = tab["fields"]
                #for field in fields:
                #    if field["type"] == "generated":
                #        continue
                        # print(field["column_title"])
            # print(tab)
        return sheet_name_array, tab

    def find_title_row(self, ws, field_name):
        #(str(ws))
        for row in ws.iter_rows():
            #print("+++++++++++++++++++++++++ after iter")
            for cell1 in row:
                if cell1.value is not None:
                    if cell1.value in field_name:
                        # print("cell="+cell1.value)
                        return cell1.row

    def put_header_fields(self, reference_objects):
        # special treatment - on report put the kupa, filename and ingestion date
        field = {"class_name": "importer.models.Kupot",
                 "field_name": "kupa",
                 "column_title": ["kupa"],
                 "type": "extracted",
                 "ref_name": "kupa"}
        self.reference_objects = self.put_in_model(
            self.reference_objects, field, None)
        field = {"class_name": "importer.models.Reports",
                 "field_name": "file_name",
                 "column_title": ["file_name"],
                 "type": "extracted",
                 "ref_name": "reports"}
        val = self.file
        self.reference_objects = self.put_in_model(
            self.reference_objects, field, val)
        field = {"class_name": "importer.models.Reports",
                 "field_name": "ingested_at",
                 "column_title": ["ingested_at"],
                 "type": "extracted",
                 "ref_name": "reports"}
        val = datetime.datetime.now()
        self.reference_objects = self.put_in_model(
            self.reference_objects, field, val)
            
    def getSheet(self, wb, sn):
            if isinstance(sn,(str)):
                return wb[sn] , sn
            for sn1 in sn:
                if sn1 in wb:
                    sn2 = sn1
                    break
            return wb[sn2] , sn2   

    def get_row(self, workbook, sheet, row):
        return next(workbook[sheet].iter_rows(min_row=row, max_row=row, values_only=False))
    
    def get_rows(self, workbook, sheet, first_row):
        return workbook[sheet].iter_rows(min_row=first_row, values_only=False)

    def get_value(self, cell):
        return cell.value
        
    def get_cell_index(self, **kwargs):
        return kwargs.get("cell").column

    def parse_first_tab(self, wb, sn, tab):
        # from first tab get report date, company, track name and track code
        # optinally get report summary as well
        self.put_header_fields(self.reference_objects)
        for field in tab["fields"]:
            if field["type"] == 'generated':
                continue
            else:
                worksheet, sn2 = self.getSheet(wb, sn)
                if worksheet is not None:    
                    for row in worksheet.iter_rows(min_row=1, max_row=4, max_col=4, values_only=False):
                        for cell in row:
                            found = False
                            if cell.value is not None:
                                if str(cell.value).startswith("{PL}PickLst"):
                                    break
                                stripped = str(cell.value).replace(
                                    '*', '').replace(":", "").strip()
                                for field1 in tab["fields"]:
                                    # in some of the reports there are multiple * characters of column titles as pointers to comments
                                    if stripped in field1["column_title"]:
                                        found = True
                                        i = 1
                                        for i in range(1, 4):
                                            val = worksheet.cell(row=cell.row, column=cell.column+i).value
                                            if val is not None:
                                                self.reference_objects = self.put_in_model(
                                                    self.reference_objects, field1, val)
                                                break
                                        break
                                    elif field1["type"] == "reference":
                                        self.reference_objects = self.put_in_model(
                                            self.reference_objects, field1, None)
                                break
                self.save_first_tab()        
               
                return sn2

    def save_first_tab(self):
        # if kupa exists
        #    compare to new
        #    warn if not equals
        #    use it
        # else use new one
        kupa = self.kupa_exists()
        if kupa is not None:
           if kupa.company != self.reference_objects["kupa"].company or kupa.track != self.reference_objects["kupa"].track:
               print("Warning: kupa not equals")
           self.reference_objects["kupa"] = kupa
           report = self.report_exists()
           if report is not None:
              if self.force:
                  self.reference_objects["reports"] = report
                  print("Warning: overwriting report")
              else:
                  raise ValueError("report already exists")
           self.reference_objects["reports"].kupa = kupa
        self.save_objects()


    def kupa_exists(self):
        dset = []
        track_no = self.reference_objects["kupa"].track_number
        try:
            dset = importer.models.Kupot.objects.filter(track_number=track_no)
        except Exception as e:
            traceback.print_exc()
        #return len(dset) > 0
        if len(dset) > 0:
            return dset[0]
        else:
            return None    

    def report_exists(self):
        dset = []
        company = self.reference_objects["kupa"].company
        track_no = self.reference_objects["kupa"].track_number
        rep_date = self.reference_objects["reports"].report_date
        dset = importer.models.Reports.objects.filter(report_date=rep_date,
                             kupa__company=company, kupa__track_number=track_no)
        if len(dset) > 0:
            return dset[0]
        else:
            return None                                          

    def put_in_model(self, objects, field, value):
        o = obj = None
        try:
            # get object from map
            if field["ref_name"] in objects:
                obj = objects[field["ref_name"]]
            if obj is not None and str(type(obj)) == "<class '"+field["class_name"]+"'>":
                o = obj
            # if not instantiated do so    
            if o is None:
                o = eval(field["class_name"]+"()")
                objects[field["ref_name"]] = o
            # if field type is reference then get the value from object map    
            if field["type"] == "reference":
                value = objects[field["field_name"]]
            if value is not None and value != "None" and str(value).strip() != '':
                # special case assumes all date fields have "date" in field name
                if ("date" in field["field_name"]):
                    value = parser.parse(value).date()
                # set the value to the right member of object    
                setattr(o, field["field_name"], value)
        except Exception as e:
            traceback.print_exc()
        return objects

    def parse_spreadsheet(self, wb):
        sheet_name_array, tab = self.find_sn(wb)
        for sh in sheet_name_array:
            # print("parsing {}".format(sh))
            titles = list()
            column_idxs = list()
            column_list = []
            title_row = 0
            found = 0
            # find title row in tab
            for field in tab["fields"]:
                if field["type"] == 'generated':
                    continue
                if field["type"] == 'extracted':
                    if title_row == 0:
                        ws, sn2=self.getSheet(wb,sh)
                        tr = self.find_title_row(ws, field["column_title"])
                        if tr is not None:
                            title_row = tr
                            break
            # 
            # create an array of fields by finding the field title in title row
            # create an array of indexes of the fields in the worksheet
            title_row = self.get_row(wb, sh, tr)
            for cell in title_row:
                if cell == '':
                    break
                cell_value = self.get_value(cell)
                if cell_value is None:
                    continue
                cell_index = self.get_cell_index(cell=cell, title_row=title_row) #title_row.index(cell)
                if cell_value is not None:
                    if str(cell_value).startswith("{PL}PickLst"):
                        break
                # in some of the reports there are multiple * characters of column titles as pointers to comments
                stripped = str(cell_value).replace('*', '').strip()
                for field in tab["fields"]:
                    found1 = False
                    if found < 2:
                        if field["type"] == 'reference':
                           found += 1
                           column_idxs.append(found)
                           column_list.append(field)
                        if field["field_name"] == "category":
                           found += 1
                           column_idxs.append(found)
                           column_list.append(field)
                    if stripped in field["column_title"]:
                        found1 = True
                        column_idxs.append(cell_index)
                        column_list.append(field)
                        break
                if not found1:
                    # field not found in mapping - log
                    uf = importer.models.UnmappedFields()
                    uf.file_name = self.file
                    uf.tab_name = sh
                    uf.field = str(cell_value)
                    uf.save()
            done = False
            # For each row in the worksheet find the values by using the index list
            # If value is valid add the value and field to the objects in the model
            for row in self.get_rows(wb, sh, tr+1):
                skip = False
                # clear details object as each row is a new details record
                if "details" in self.reference_objects:
                    del self.reference_objects["details"]
                # put data in details object
                for i in range(len(column_idxs)):
                    field = column_list[i]
                    if field["type"] == 'reference':
                        value = None
                    elif field["field_name"] == "category":
                        value = sh
                    else:
                        ind = column_idxs[i]
                        cell1 = row[ind-1]
                        value = str(self.get_value(cell1))
                         # special treatment - stock_name must be populated or row is not a data row
                        if "stock_name" == field["field_name"] and (value is None or value == 'None' or value == ''):
                            skip = True
                            break
                        #if cell1 is None or cell1 == 'None' or cell1 == '':
                        #    break
                        # if value is a formula - currently only relevant in pyxl
                        if value.startswith("="):
                            try:
                                # field is a formula field - calculate it
                                value = self.interface.calc_cell(
                                    cell1.coordinate, sh)
                                # value = self.excel.evaluate("'"+sh+"'!"+cell1.coordinate)
                            except Exception as e:
                                traceback.print_exc()
                                fni = importer.models.FilesNotIngested()
                                fni.file_name = self.file
                                fni.info = sh+"-"+cell1.coordinate + \
                                    "\n\r+++"+str(e)
                                fni.save()
                                break
                            # apply format - is it really needed?
                            f_format = cell1.number_format
                            if f_format == "0.00%":
                                # match f_format:
                                #    case "0.00%":
                                value = value*100
                            #    case "#,##0.00" :
                            #        value = '{:.2}'.format(value)
                            # elif f_format == "#,##0.00" :
                            # round to 2 decimal points
                            value = f"{value:.2f}"
                            # print(self.file+","+sh+ "-"+cell1.coordinate+"-"+str(value) +" format ="+f_format )
                    # special treatment - row starting with * signals end of data
                    if '*' in str(value):
                        done = True
                    else:
                        self.reference_objects = self.put_in_model(
                            self.reference_objects, field, value)
                    if done:
                        break
                if not skip:
                    self.save_objects()
                if done:
                    break

    def save_objects(self):
        for o in self.reference_objects.values():
            try:
                o.save()
            except Exception as e:
                traceback.print_exc()
                fni = importer.models.FilesNotIngested()
                fni.file_name = self.file
                fni.info = type(o).__name__ +\
                    "\n\r+++"+str(e)
                fni.save()

    def ingest(self, filename, file_stream):
        try:
            self.file = filename
            wb = load_workbook(file_stream)
            # OpenpyxlInterface allows calculating formulas
            self.interface = OpenpyxlInterface(wb=wb, use_cache=True)
            # self.excel = ExcelCompiler(excel=wb)
            # wb.calculation.calcOnLoad = True
            self.parse_spreadsheet(wb)
            self.reference_objects.clear()
            return True
        except ValueError: 

            traceback.print_exc()
            print("report already exists")
            return False

        except Exception as e:
            # log failed files
            traceback.print_exc()
            fni = importer.models.FilesNotIngested()
            fni.file_name = filename
            fni.info = "Failed to read workbook\n\r"+str(e)
            fni.save()
            return False


def xls_import(path, file):
    if path is None:
        ingester = xls_ingester()
        ingester.ingest(file)
    else:
        files = os.listdir(path)
        files = [os.path.join(path, f) for f in files if not f.startswith(".")]
        for f in files:
            ingester = xls_ingester()
            ingester.ingest(f)


def main():  # Main
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--full_file_name', type=str)
    parser.add_argument('-d', '--path', type=str)
    args = parser.parse_args()
    if args.path is None and args.full_file_name is None:
        print("specify either single file or directory")
        return
    if args.path is not None and args.full_file_name is not None:
        print("specify either single file or directory")
        return
    xls_import(args.path, args.full_file_name)


if __name__ == '__main__':
    main()
